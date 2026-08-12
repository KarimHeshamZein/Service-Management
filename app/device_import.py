"""Excel template and validation helpers for technician service entry."""
from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass, field
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .models import InstalledDevice, PricingItem, WorkSite


DEVICE_IMPORT_HEADERS = (
    "Item / Device Name",
    "Model",
    "Serial Number",
    "IMEI",
    "SIM Serial Number",
    "Sim Type",
    "Main Project",
    "Sub Project",
    "Site",
    "Remarks",
    "Status",
)
SIM_TYPES = {"zain": "Zain", "mobily": "Mobily", "stc": "STC"}
MAX_IMPORT_ROWS = 1000
MAX_XLSX_BYTES = 5 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024


@dataclass(slots=True)
class DeviceImportRow:
    row_number: int
    item_name: str
    model: str
    serial_number: str
    imei: str = ""
    iccid: str = ""
    sim_type: str = ""
    main_project: str = ""
    sub_project: str = ""
    site: str = ""
    remarks: str = ""
    status: str = ""
    pricing_item_id: int | None = None
    device_id: int | None = None
    site_id: int | None = None
    installed_device_id: int | None = None
    asset_conflicts: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def token_payload(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "item_name": self.item_name,
            "model": self.model,
            "serial_number": self.serial_number,
            "imei": self.imei or None,
            "iccid": self.iccid or None,
            "sim_type": self.sim_type or None,
            "main_project": self.main_project,
            "sub_project": self.sub_project,
            "site": self.site,
            "remarks": self.remarks or None,
            "status": self.status,
            "pricing_item_id": self.pricing_item_id,
            "device_id": self.device_id,
            "site_id": self.site_id,
            "installed_device_id": self.installed_device_id,
            "asset_conflicts": self.asset_conflicts,
        }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def build_device_template(
    items: Iterable[PricingItem],
    sites: Iterable[WorkSite],
    *,
    entry_label: str,
    main_project_names: Iterable[str] = (),
    sub_project_names: Iterable[str] = (),
) -> bytes:
    """Return a polished technician entry template."""
    item_rows = sorted({(item.name, item.model) for item in items})
    site_names = sorted({site.name for site in sites if site.is_active})
    main_names = sorted({name for name in main_project_names if name})
    sub_names = sorted({name for name in sub_project_names if name})

    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    sheet = workbook.active
    sheet.title = "Device Data"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(DEVICE_IMPORT_HEADERS))}201"
    sheet.append(DEVICE_IMPORT_HEADERS)
    for _ in range(200):
        sheet.append([""] * len(DEVICE_IMPORT_HEADERS))

    navy, blue, pale = "17324D", "1F6F8B", "EAF3F6"
    thin = Side(style="thin", color="D7E0E6")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=blue))
    sheet.row_dimensions[1].height = 36
    widths = (28, 20, 22, 18, 24, 15, 24, 24, 22, 34, 14)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2, max_row=201):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if cell.row % 2 else pale)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top")
            cell.number_format = "@"
    for row_number in range(2, 202):
        sheet.cell(
            row_number,
            11,
            f'=IF(COUNTA(A{row_number}:J{row_number})=0,"",IF(COUNTA(A{row_number}:J{row_number})=10,"Valid","Invalid"))',
        )
    sheet.conditional_formatting.add(
        "K2:K201",
        FormulaRule(
            formula=['$K2="Valid"'],
            fill=PatternFill("solid", fgColor="E8F7EF"),
            font=Font(color="17633A", bold=True),
        ),
    )
    sheet.conditional_formatting.add(
        "K2:K201",
        FormulaRule(
            formula=['$K2="Invalid"'],
            fill=PatternFill("solid", fgColor="FDECEC"),
            font=Font(color="A12622", bold=True),
        ),
    )

    instructions = workbook.create_sheet("Instructions")
    instructions.sheet_view.showGridLines = False
    instructions.merge_cells("A1:F1")
    instructions["A1"] = f"{entry_label} - Device Data"
    instructions["A1"].fill = PatternFill("solid", fgColor=navy)
    instructions["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    instructions["A1"].alignment = Alignment(vertical="center")
    instructions.row_dimensions[1].height = 34
    instructions["A3"] = "How to use this template"
    instructions["A3"].font = Font(color=navy, bold=True, size=12)
    guidance = (
        "1. Enter one device per row on the Device Data sheet.",
        "2. Enter the device data in columns A-J. Do not type in the Status column.",
        "3. Use the dropdown values where available and do not rename or reorder columns.",
        "4. IMEI and SIM Serial Number are stored as entered.",
        "5. Sim Type may be blank, Zain, Mobily or STC.",
        "6. Status is calculated automatically: Valid when columns A-J are complete, otherwise Invalid.",
        "7. Upload the completed workbook in Data Entry and review the preview before submitting.",
    )
    for row_number, line in enumerate(guidance, 5):
        instructions.cell(row_number, 1, line)
    instructions.column_dimensions["A"].width = 110

    lists = workbook.create_sheet("Lists")
    lists.append(("Item Names", "Models", "SIM Types", "Main Projects", "Sub Projects", "Sites"))
    max_rows = max(len(item_rows), len(SIM_TYPES), len(main_names), len(sub_names), len(site_names), 1)
    sim_values = list(SIM_TYPES.values())
    for index in range(max_rows):
        item = item_rows[index] if index < len(item_rows) else ("", "")
        lists.append(
            (
                *item,
                sim_values[index] if index < len(sim_values) else "",
                main_names[index] if index < len(main_names) else "",
                sub_names[index] if index < len(sub_names) else "",
                site_names[index] if index < len(site_names) else "",
            )
        )
    lists.sheet_state = "hidden"

    validations = {
        "A": ("A", len(item_rows), False),
        "B": ("B", len(item_rows), False),
        "F": ("C", len(sim_values), True),
        "G": ("D", len(main_names), False),
        "H": ("E", len(sub_names), False),
        "I": ("F", len(site_names), False),
    }
    for target, (source, length, allow_blank) in validations.items():
        if length < 1:
            continue
        validator = DataValidation(
            type="list",
            formula1=f"'Lists'!${source}$2:${source}${length + 1}",
            allow_blank=allow_blank,
        )
        validator.error = "Choose a value from the approved list."
        validator.errorTitle = "Invalid selection"
        validator.prompt = "Select an approved value."
        validator.promptTitle = f"{entry_label} device data"
        validator.showErrorMessage = True
        validator.showInputMessage = True
        sheet.add_data_validation(validator)
        validator.add(f"{target}2:{target}201")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def validate_device_workbook(
    data: bytes,
    *,
    filename: str,
    db: Session,
    items: Iterable[PricingItem],
    project_id: int,
    site: WorkSite,
    entry_kind: str,
) -> tuple[list[DeviceImportRow], list[str]]:
    """Validate a technician workbook without mutating records or assets."""
    if not filename.lower().endswith(".xlsx"):
        return [], ["Upload an Excel .xlsx file created from the downloaded template."]
    if not data:
        return [], ["The uploaded Excel file is empty."]
    if len(data) > MAX_XLSX_BYTES:
        return [], ["The Excel file is larger than the 5 MB import limit."]
    if not data.startswith(b"PK"):
        return [], ["The uploaded file is not a valid Excel .xlsx workbook."]
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            if sum(entry.file_size for entry in archive.infolist()) > MAX_XLSX_UNCOMPRESSED_BYTES:
                return [], ["The Excel workbook expands beyond the safe processing limit."]
    except (OSError, zipfile.BadZipFile):
        return [], ["The uploaded file is not a valid Excel .xlsx workbook."]
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return [], ["The Excel workbook could not be read. Download a new template and try again."]
    if "Device Data" not in workbook.sheetnames:
        return [], ["The workbook is missing the Device Data sheet."]
    sheet = workbook["Device Data"]
    headers = tuple(_cell_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
    if headers[: len(DEVICE_IMPORT_HEADERS)] != DEVICE_IMPORT_HEADERS:
        return [], ["The Device Data columns were changed. Download a new template and keep the headers unchanged."]

    item_map = {(item.name.casefold(), item.model.casefold()): item for item in items}
    existing = list(
        db.scalars(
            select(InstalledDevice)
            .options(selectinload(InstalledDevice.work_site_evidence))
            .where(InstalledDevice.is_active.is_(True))
        )
    )
    by_serial = {
        device.serial_number.casefold(): device
        for device in existing
        if device.serial_number
    }
    by_imei = {device.imei: device for device in existing if device.imei}
    by_iccid = {device.iccid: device for device in existing if device.iccid}
    rows: list[DeviceImportRow] = []
    structural_errors: list[str] = []
    for row_number, cells in enumerate(
        sheet.iter_rows(min_row=2, max_col=len(DEVICE_IMPORT_HEADERS), values_only=True), 2
    ):
        values = [_cell_text(value) for value in cells]
        if not any(values):
            continue
        if len(rows) >= MAX_IMPORT_ROWS:
            structural_errors.append(f"Import at most {MAX_IMPORT_ROWS} device rows at a time.")
            break
        row = DeviceImportRow(row_number, *values)
        for value, message in (
            (row.item_name, "Item / Device Name is required."),
            (row.model, "Model is required."),
            (row.serial_number, "Serial Number is required."),
            (row.site, "Site is required."),
        ):
            if not value:
                row.errors.append(message)

        item = item_map.get((row.item_name.casefold(), row.model.casefold()))
        if row.item_name and row.model and item is None:
            row.errors.append("The Item / Device Name and Model do not match an available Items Library entry.")
        elif item is not None and item.legacy_device is not None and item.legacy_device.is_active:
            row.pricing_item_id = item.id
            row.device_id = item.legacy_device.id
        elif item is not None:
            row.errors.append("This item is not available for service entry.")

        row.site_id = site.id
        # Status is an Excel completeness indicator, not a service result.
        # Recompute it here so dashboard preview never depends on a formula cache
        # or on a user-overwritten Status cell.
        row.status = "Valid" if all(values[:10]) else "Invalid"
        if row.sim_type:
            normalized = row.sim_type.casefold()
            if normalized not in SIM_TYPES:
                row.errors.append("Sim Type must be blank, Zain, Mobily or STC.")
            else:
                row.sim_type = normalized
        if row.iccid and not row.sim_type:
            row.warnings.append("SIM Serial Number is present but Sim Type is blank.")
        if row.sim_type and not row.iccid:
            row.warnings.append("Sim Type is present but SIM Serial Number is blank.")

        matched = by_serial.get(row.serial_number.casefold()) if row.serial_number else None
        if entry_kind == "installation":
            if matched:
                row.errors.append("Serial Number is already registered as an Installed Asset.")
        elif matched:
            if matched.site_id != project_id or matched.effective_work_site_id != site.id:
                row.errors.append("Serial Number belongs to an Installed Asset outside the selected Project and Site.")
            elif row.device_id and matched.device_id != row.device_id:
                row.errors.append("Serial Number belongs to a different Item / Device Name and Model.")
            else:
                row.installed_device_id = matched.id
                for attribute, label in (
                    ("imei", "IMEI"),
                    ("iccid", "SIM Serial Number"),
                    ("sim_type", "Sim Type"),
                    ("remarks", "Remarks"),
                ):
                    incoming = getattr(row, attribute)
                    current = getattr(matched, attribute)
                    if incoming and current and incoming != current:
                        row.asset_conflicts.append(label)
                if row.asset_conflicts:
                    row.warnings.append(
                        "Existing asset data differs for: " + ", ".join(row.asset_conflicts) + ". Confirm replacement before saving."
                    )
        elif entry_kind != "installation":
            row.warnings.append("No matching Installed Asset was found; data will be stored on this maintenance record only.")

        if row.imei and row.imei in by_imei and by_imei[row.imei] is not matched:
            row.errors.append("IMEI is already registered to another Installed Asset.")
        if row.iccid and row.iccid in by_iccid and by_iccid[row.iccid] is not matched:
            row.errors.append("SIM Serial Number is already registered to another Installed Asset.")
        rows.append(row)

    if not rows and not structural_errors:
        structural_errors.append("The Device Data sheet does not contain any device rows.")
        return rows, structural_errors

    for attribute, label in (("serial_number", "Serial Number"), ("imei", "IMEI"), ("iccid", "SIM Serial Number")):
        values = [getattr(row, attribute).casefold() for row in rows if getattr(row, attribute)]
        duplicates = {value for value in values if values.count(value) > 1}
        for row in rows:
            if getattr(row, attribute) and getattr(row, attribute).casefold() in duplicates:
                row.errors.append(f"Duplicate {label} in the uploaded workbook.")
    return rows, structural_errors
