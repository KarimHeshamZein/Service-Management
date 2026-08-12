from __future__ import annotations

import io
import re
from datetime import timedelta

from openpyxl import load_workbook
from pypdf import PdfReader

from app.device_import import DEVICE_IMPORT_HEADERS, build_device_template
from app.helpers import to_display
from app.models import (
    InstalledDevice,
    InstalledDeviceSite,
    GeneralMaintenanceRecord,
    InstallationRecord,
    MaintenanceRecord,
    PricingItem,
    ServiceReport,
    SubProject,
    SubProjectSite,
    WorkSite,
)
from app.pdf_text import pdf_text, style_for_pdf_text
from app.routers.structured_reports import _record_tree
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
from tests.conftest import (
    ADMIN,
    CUSTOMER_A,
    CUSTOMER_B,
    LEADER_A,
    csrf_of,
    ensure_service_quotation,
    login,
    logout,
    make_image,
    submit_installation,
    submit_record,
)


def test_report_tree_expands_one_record_across_two_sub_projects():
    record = {
        "id": 77,
        "project_id": 3,
        "customer_name": "Newtest Project",
        "sub_project_id": 4,
        "sub_project_name": "Sub-Sheraton",
        "work_site_id": 1,
        "work_site_name": "gate 1",
        "site_name": "gate 1",
        "site_sections": [
            {
                "project_id": 3,
                "project_name": "Newtest Project",
                "sub_project_id": 4,
                "sub_project_name": "Sub-Sheraton",
                "work_site_id": 1,
                "work_site_name": "gate 1",
            },
            {
                "project_id": 3,
                "project_name": "Newtest Project",
                "sub_project_id": 5,
                "sub_project_name": "Sub-Sub-Sheraton",
                "work_site_id": 2,
                "work_site_name": "gate 2",
            },
        ],
    }

    tree = _record_tree([record], {})

    assert tree[0]["record_count"] == 1
    assert [sub["name"] for sub in tree[0]["sub_projects"]] == [
        "Sub-Sheraton",
        "Sub-Sub-Sheraton",
    ]
    assert [sub["sites"][0]["name"] for sub in tree[0]["sub_projects"]] == [
        "gate 1",
        "gate 2",
    ]
    assert all(
        sub["sites"][0]["records"][0]["id"] == 77
        for sub in tree[0]["sub_projects"]
    )


def _record_id(db) -> int:
    return db.query(InstallationRecord).order_by(InstallationRecord.id.desc()).first().id


def _report_payload(client, record_id: int, **extra):
    payload = {
        "csrf_token": csrf_of(client, "/reports/installation/new"),
        "name": "Gate camera commissioning",
        "report_date": "2026-08-09",
        "team_leader_id": "2",
        "technician_ids": "3",
        "record_ids": str(record_id),
        "notes": "Customer-facing installation handover.",
    }
    payload.update(extra)
    return payload


def _device_workbook(
    db,
    *,
    serial: str,
    remarks: str = "Field data",
    imei: str = "490154203237518",
    iccid: str = "899660123456789012",
) -> bytes:
    content = build_device_template(
        [db.query(PricingItem).one()],
        [db.get(WorkSite, 1)],
        entry_label="Service Entry",
        main_project_names=["Tower A"],
        sub_project_names=["General"],
    )
    workbook = load_workbook(io.BytesIO(content))
    values = (
        "IP Camera",
        "P3265-LV",
        serial,
        imei,
        iccid,
        "STC",
        "Tower A",
        "General",
        "Gate 1",
        remarks,
    )
    for column, value in enumerate(values, 1):
        workbook["Device Data"].cell(row=2, column=column, value=value)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_pdf_text_removes_invisible_item_name_characters_and_centers_arabic():
    assert pdf_text("Sticker Speed \u200b\u200bsign 160*120") == (
        "Sticker Speed sign 160*120"
    )
    centered = ParagraphStyle("CenteredCaption", alignment=TA_CENTER)
    assert style_for_pdf_text("وصف الصورة", centered).alignment == TA_CENTER


def test_admin_can_delete_report_only_without_deleting_source_record(client, db):
    login(client, *LEADER_A)
    assert submit_installation(client, serial_number="REPORT-ONLY-DELETE").status_code == 303
    record_id = _record_id(db)
    created = client.post(
        "/reports/installation",
        data=_report_payload(client, record_id),
    )
    assert created.status_code == 303
    report_id = db.query(ServiceReport).one().id
    asset = db.get(InstallationRecord, record_id).installed_device
    asset_id = asset.id
    asset.source_report_id = report_id
    db.commit()

    logout(client)
    login(client, *ADMIN)
    token = csrf_of(client, f"/reports/installation/{report_id}")
    deleted = client.post(
        f"/reports/installation/{report_id}/delete",
        data={"csrf_token": token},
    )

    assert deleted.status_code == 303
    db.expire_all()
    assert db.get(ServiceReport, report_id) is None
    assert db.get(InstallationRecord, record_id) is not None
    preserved_asset = db.get(InstalledDevice, asset_id)
    assert preserved_asset is not None
    assert preserved_asset.source_report_id is None


def test_deleting_source_record_deletes_entire_linked_report(client, db):
    login(client, *LEADER_A)
    assert submit_installation(client, serial_number="REPORT-CASCADE-A").status_code == 303
    first_id = _record_id(db)
    assert submit_installation(client, serial_number="REPORT-CASCADE-B").status_code == 303
    second_id = _record_id(db)
    created = client.post(
        "/reports/installation",
        data=_report_payload(client, first_id, record_ids=[str(first_id), str(second_id)]),
    )
    assert created.status_code == 303
    report_id = db.query(ServiceReport).one().id

    logout(client)
    login(client, *ADMIN)
    detail = client.get(f"/installations/records/{first_id}")
    assert "appears in 1 generated report(s)" in detail.text
    token = csrf_of(client, f"/installations/records/{first_id}")
    deleted = client.post(
        f"/installations/records/{first_id}/delete",
        data={"csrf_token": token},
    )

    assert deleted.status_code == 303
    db.expire_all()
    assert db.get(InstallationRecord, first_id) is None
    assert db.get(InstallationRecord, second_id) is not None
    assert db.get(ServiceReport, report_id) is None


def test_saved_installation_report_has_fixed_creator_and_customer_scope(client, db):
    login(client, *LEADER_A)
    assert submit_installation(
        client,
        handover_notes="Customer accepted the commissioned camera.",
        photos=[
            ("before_photos_0", ("before.jpg", make_image((80, 20, 20)), "image/jpeg")),
            ("after_photos_0", ("after.jpg", make_image((20, 80, 20)), "image/jpeg")),
        ],
    ).status_code == 303
    record_id = _record_id(db)

    response = client.post(
        "/reports/installation",
        data={**_report_payload(client, record_id), "created_by_id": "1"},
    )
    assert response.status_code == 303
    db.expire_all()
    report = db.query(ServiceReport).one()
    assert report.report_number.startswith("IR-2026-")
    assert report.created_by_id == 2
    assert report.created_by_name == "Leader One"
    assert report.team_leader_id == 2
    assert [member.user_id for member in report.technicians] == [3]
    assert report.record_links[0].installation_record_id == record_id
    assert report.record_links[0].sub_project_name == "General"

    detail = client.get(response.headers["location"])
    assert detail.status_code == 200
    assert "Gate camera commissioning" in detail.text
    assert detail.text.index("PDF Preview &amp; Download") < detail.text.index("Report Information")
    pdf = client.get(f"/reports/installation/{report.id}/pdf")
    assert pdf.status_code == 200
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content.startswith(b"%PDF")
    pdf_pages = [page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf.content)).pages]
    pdf_text = "\n".join(pdf_pages)
    assert "installation.jpg" not in pdf_text
    assert "Installation notes" in pdf_text
    assert "Mounted, connected, configured and commissioned the equipment." in pdf_text
    assert "Handover notes" in pdf_text
    assert "Customer accepted the commissioned camera." in pdf_text
    assert "APPROVALS" in pdf_text
    assert "Customer Representative" in pdf_text
    assert "Afaqy Representative" in pdf_text
    assert "Project Manager" in pdf_text
    assert pdf_text.count("Signature & Stamp") == 3
    detail_page = next(index for index, text in enumerate(pdf_pages) if "Installation notes" in text)
    evidence_page = next(index for index, text in enumerate(pdf_pages) if "Before Installation" in text)
    assert evidence_page > detail_page
    assert "After Installation" in pdf_text
    manual_pdf = client.get(f"/reports/installation/{report.id}/pdf?include_device_data=true")
    manual_text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(manual_pdf.content)).pages)
    assert "Device Data" not in manual_text

    logout(client)
    login(client, *CUSTOMER_A)
    assert client.get(f"/reports/installation/{report.id}").status_code == 200
    logout(client)
    login(client, *CUSTOMER_B)
    assert client.get(f"/reports/installation/{report.id}").status_code == 404


def test_entry_excel_download_preview_and_installation_asset_creation(client, db):
    general = SubProject(project_id=1, name="General")
    general.site_assignments = [SubProjectSite(site_id=1)]
    db.add(general)
    db.commit()
    login(client, *LEADER_A)
    download = client.get("/data-entry/installation/device-template")
    assert download.status_code == 200
    assert download.content.startswith(b"PK")
    assert 'filename="installation-device-data-template.xlsx"' in download.headers["content-disposition"]
    item = db.query(PricingItem).one()
    content = build_device_template(
        [item],
        [db.get(WorkSite, 1)],
        entry_label="Installation",
        main_project_names=["Tower A"],
        sub_project_names=["General"],
    )
    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames == ["Device Data", "Instructions", "Lists"]
    assert workbook["Lists"].sheet_state == "hidden"
    sheet = workbook["Device Data"]
    assert tuple(cell.value for cell in sheet[1]) == DEVICE_IMPORT_HEADERS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref.endswith("201")
    assert len(sheet.data_validations.dataValidation) == 6
    assert '"Valid"' in sheet["K2"].value
    assert '"Invalid"' in sheet["K2"].value
    assert len(sheet.conditional_formatting) == 1
    assert sheet["A1"].fill.fgColor.rgb.endswith("17324D")
    values = (
        "IP Camera",
        "P3265-LV",
        "ENTRY-ASSET-001",
        "12345",
        "SIM-ABC",
        "STC",
        "Tower A",
        "General",
        "North gate cabinet",
        "North gate camera",
    )
    for column, value in enumerate(values, 1):
        sheet.cell(row=2, column=column, value=value)
    output = io.BytesIO()
    workbook.save(output)

    csrf, form_token = (
        re.search(r'name="csrf_token" value="([^"]+)"', (page := client.get("/installations/submit")).text).group(1),
        re.search(r'name="form_token" value="([^"]+)"', page.text).group(1),
    )
    preview = client.post(
        "/data-entry/installation/device-import-preview",
        data={
            "csrf_token": csrf,
            "project_id": "1",
            "sub_project_id": str(general.id),
            "work_site_id": "1",
        },
        files={
            "device_file": (
                "completed.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 200
    preview_payload = preview.json()
    assert preview_payload["ok"] is True
    assert preview_payload["rows"][0]["status"] == "Valid"

    created = submit_installation(
        client,
        serial_number="ENTRY-ASSET-001",
        sub_project_id=str(general.id),
        tokens=(csrf, form_token),
        device_import_token=preview_payload["token"],
    )
    assert created.status_code == 303
    db.expire_all()
    asset = db.query(InstalledDevice).filter_by(serial_number="ENTRY-ASSET-001").one()
    assert asset.imei == "12345"
    assert asset.iccid == "SIM-ABC"
    assert asset.sim_type == "stc"
    assert asset.phone_number is None
    assert asset.remarks == "North gate camera"
    record = db.query(InstallationRecord).filter_by(serial_number="ENTRY-ASSET-001").one()
    assert record.work_items[0].imei == "12345"
    assert record.work_items[0].imported_from_excel is True
    assert record.work_items[0].location_name == "North gate cabinet"
    report_response = client.post(
        "/reports/installation",
        data=_report_payload(client, record.id, include_device_data="1"),
    )
    assert report_response.status_code == 303
    report = db.query(ServiceReport).one()
    pdf = client.get(f"/reports/installation/{report.id}/pdf?include_device_data=true")
    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")


def test_report_creation_rejects_unknown_record(client):
    login(client, *LEADER_A)
    response = client.post(
        "/reports/installation",
        data=_report_payload(client, 999999),
    )
    assert response.status_code == 422
    assert "unavailable or outside" in response.text


def test_report_includes_browser_entered_site_table(client, db):
    login(client, *LEADER_A)
    created = submit_installation(
        client,
        serial_number="",
        data_scope_index="0",
        data_item_name="Entrance Camera",
        data_model="P3265-LV",
        data_serial_number="TABLE-SN-001",
        data_remarks="North entrance",
    )
    assert created.status_code == 303
    record = db.query(InstallationRecord).one()
    saved = client.post(
        "/reports/installation",
        data=_report_payload(client, record.id, include_device_data="1"),
    )
    assert saved.status_code == 303
    report = db.query(ServiceReport).one()
    response = client.get(
        f"/reports/installation/{report.id}/pdf?include_device_data=true"
    )
    assert response.status_code == 200
    text = "\n".join(
        page.extract_text() or ""
        for page in PdfReader(io.BytesIO(response.content)).pages
    )
    assert "SERVICE DATA TABLES" in text
    assert "Entrance Camera" in text
    assert "TABLE-SN-001" in text
    assert "SITE | Gate 1" in text


def test_installation_photo_descriptions_and_order_are_saved(client, db):
    login(client, *LEADER_A)
    response = submit_installation(
        client,
        serial_number="PHOTO-DESCRIPTIONS-001",
        photos=[
            ("before_photos_0", ("first.jpg", make_image((10, 20, 30)), "image/jpeg")),
            ("before_photos_0", ("second.jpg", make_image((30, 20, 10)), "image/jpeg")),
        ],
        before_photo_descriptions_0=["Entrance before work", "Cable route before work"],
    )
    assert response.status_code == 303
    db.expire_all()
    record = db.query(InstallationRecord).order_by(InstallationRecord.id.desc()).first()
    assert [photo.description for photo in record.work_items[0].photos] == [
        "Entrance before work",
        "Cable route before work",
    ]
    assert [photo.position for photo in record.work_items[0].photos] == [0, 1]


def test_entry_scope_saves_selected_sub_project(client, db):
    sub_project = SubProject(project_id=1, name="Security Systems")
    sub_project.site_assignments = [SubProjectSite(site_id=1)]
    db.add(sub_project)
    db.commit()
    login(client, *LEADER_A)
    page = client.get("/installations/submit")
    assert page.status_code == 200
    assert "Security Systems" in page.text
    response = submit_installation(
        client,
        serial_number="SUB-PROJECT-001",
        sub_project_id=str(sub_project.id),
    )
    assert response.status_code == 303
    db.expire_all()
    record = db.query(InstallationRecord).order_by(InstallationRecord.id.desc()).first()
    assert record.sub_project_id == sub_project.id
    assert record.sub_project_name == "Security Systems"
    assert record.installed_device.sub_project_id == sub_project.id


def test_saved_report_workflow_is_shared_by_both_maintenance_types(client, db):
    login(client, *LEADER_A)
    assert submit_record(
        client,
        notes="Preventive notes for PDF.",
        issue_description="Preventive issue for PDF.",
        recommendations="Preventive recommendation for PDF.",
    ).status_code == 303
    preventive = db.query(MaintenanceRecord).one()
    preventive_response = client.post(
        "/reports/preventive-maintenance",
        data={
            **_report_payload(client, preventive.id),
            "name": "Preventive visit",
        },
    )
    assert preventive_response.status_code == 303

    token_page = client.get("/general-maintenance")
    csrf = re.search(r'name="csrf_token" value="([^"]+)"', token_page.text).group(1)
    form_token = re.search(r'name="form_token" value="([^"]+)"', token_page.text).group(1)
    general_response = client.post(
        "/general-maintenance/submit",
        data={
            "csrf_token": csrf,
            "form_token": form_token,
            "project_id": "1",
            "work_site_id": "1",
            "service_type_id": "1",
            "installed_device_id": "1",
            "result_0": "completed_successfully",
            "notes": "Completed corrective maintenance.",
            "issue_description": "Corrective issue for PDF.",
            "recommendations": "Corrective recommendation for PDF.",
        },
        files={"photos_0": ("proof.jpg", make_image(), "image/jpeg")},
    )
    assert general_response.status_code == 303
    db.expire_all()
    general = db.query(GeneralMaintenanceRecord).one()
    saved = client.post(
        "/reports/maintenance",
        data={**_report_payload(client, general.id), "name": "Maintenance visit"},
    )
    assert saved.status_code == 303
    db.expire_all()
    reports = db.query(ServiceReport).order_by(ServiceReport.id).all()
    assert reports[0].record_links[0].preventive_record_id == preventive.id
    assert reports[1].record_links[0].maintenance_record_id == general.id
    preventive_pdf = client.get(
        f"/reports/preventive-maintenance/{reports[0].id}/pdf"
    )
    maintenance_pdf = client.get(f"/reports/maintenance/{reports[1].id}/pdf")
    preventive_text = "\n".join(
        page.extract_text() or ""
        for page in PdfReader(io.BytesIO(preventive_pdf.content)).pages
    )
    maintenance_text = "\n".join(
        page.extract_text() or ""
        for page in PdfReader(io.BytesIO(maintenance_pdf.content)).pages
    )
    for label in ("Issue found", "Recommendations"):
        assert label in preventive_text
        assert label in maintenance_text
    for label in ("Maintenance notes", "Model", "Serial number"):
        assert label not in preventive_text
        assert label not in maintenance_text
    assert "Preventive issue for PDF." in preventive_text
    assert "Preventive recommendation for PDF." in preventive_text
    assert "Corrective issue for PDF." in maintenance_text
    assert "Corrective recommendation for PDF." in maintenance_text


def test_multi_device_pdf_keeps_each_device_with_its_own_evidence(client, db):
    second = InstalledDevice(
        site_id=1,
        device_id=1,
        customer_name="Tower A",
        site_name="Gate 1",
        device_name="Gate Barrier",
        manufacturer="Barrier Co",
        device_model="GB-200",
        serial_number="MULTI-PDF-002",
    )
    second.work_site_evidence = InstalledDeviceSite(site_id=1, site_name="Gate 1")
    db.add(second)
    db.commit()

    login(client, *LEADER_A)
    page = client.get("/general-maintenance")
    response = client.post(
        "/general-maintenance/submit",
        data={
            "csrf_token": re.search(r'name="csrf_token" value="([^"]+)"', page.text).group(1),
            "form_token": re.search(r'name="form_token" value="([^"]+)"', page.text).group(1),
            "project_id": "1",
            "work_site_id": "1",
            "service_type_id": ["1", "1"],
            "installed_device_id": ["1", str(second.id)],
            "result_0": "completed_successfully",
            "result_1": "completed_successfully",
            "notes": ["First device notes.", "Second device notes."],
            "issue_description": ["First issue.", "Second issue."],
            "recommendations": ["First recommendation.", "Second recommendation."],
        },
        files=[
            ("before_photos_0", ("first-before.jpg", make_image((80, 20, 20)), "image/jpeg")),
            ("after_photos_0", ("first-after.jpg", make_image((20, 80, 20)), "image/jpeg")),
            ("before_photos_1", ("second-before.jpg", make_image((20, 20, 80)), "image/jpeg")),
            ("after_photos_1", ("second-after.jpg", make_image((80, 80, 20)), "image/jpeg")),
        ],
    )
    assert response.status_code == 303
    record = db.query(GeneralMaintenanceRecord).one()
    saved = client.post(
        "/reports/maintenance",
        data={**_report_payload(client, record.id), "name": "Multi-device maintenance"},
    )
    assert saved.status_code == 303
    report = db.query(ServiceReport).one()

    pdf = client.get(f"/reports/maintenance/{report.id}/pdf")
    pages = [page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf.content)).pages]
    first_detail = next(index for index, text in enumerate(pages) if "First device notes." in text)
    first_evidence = next(
        index
        for index, text in enumerate(pages)
        if index > first_detail and "PHOTO EVIDENCE" in text and "IP Camera" in text
    )
    first_after = next(
        index for index, text in enumerate(pages) if index >= first_evidence and "After Maintenance" in text
    )
    second_detail = next(index for index, text in enumerate(pages) if "Second device notes." in text)
    second_evidence = next(
        index
        for index, text in enumerate(pages)
        if index > second_detail and "PHOTO EVIDENCE" in text and "Gate Barrier" in text
    )
    second_after = next(
        index for index, text in enumerate(pages) if index >= second_evidence and "After Maintenance" in text
    )

    assert first_detail < first_evidence <= first_after < second_detail < second_evidence <= second_after
    assert "Second device notes." not in pages[first_detail]
    assert "First device notes." not in pages[second_detail]
    assert f"Record {record.record_number}" in pages[first_detail]
    assert f"Record {record.record_number}" in pages[second_detail]
    assert sum(text.count(f"Record {record.record_number}") for text in pages) == 2
    assert "MAIN PROJECT" in pages[first_detail]
    assert "SUB PROJECT" in pages[first_detail]
    assert "SITE" in pages[first_detail]


def test_preventive_excel_requires_confirmation_before_replacing_asset_values(client, db):
    general = SubProject(project_id=1, name="General")
    general.site_assignments = [SubProjectSite(site_id=1)]
    asset = db.get(InstalledDevice, 1)
    asset.phone_number = "0500000000"
    asset.remarks = "Old remarks"
    db.add(general)
    db.commit()
    login(client, *LEADER_A)
    csrf = csrf_of(client, "/maintenance/submit")
    preview = client.post(
        "/data-entry/preventive-maintenance/device-import-preview",
        data={
            "csrf_token": csrf,
            "project_id": "1",
            "sub_project_id": str(general.id),
            "work_site_id": "1",
        },
        files={"device_file": ("completed.xlsx", _device_workbook(db, serial="BASE-SN-001", remarks="Updated remarks"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    payload = preview.json()
    assert payload["has_asset_conflicts"] is True

    rejected = submit_record(
        client,
        installed_device_id="catalog:1",
        sub_project_id=str(general.id),
        device_import_token=payload["token"],
    )
    assert rejected.status_code == 422
    accepted = submit_record(
        client,
        installed_device_id="catalog:1",
        sub_project_id=str(general.id),
        device_import_token=payload["token"],
        confirm_asset_overwrites="1",
    )
    assert accepted.status_code == 303
    db.expire_all()
    asset = db.get(InstalledDevice, 1)
    assert asset.phone_number == "0500000000"
    assert asset.remarks == "Updated remarks"
    item = db.query(MaintenanceRecord).one().work_items[0]
    assert item.installed_device_id == asset.id
    assert item.imei == "490154203237518"
    assert item.iccid == "899660123456789012"
    assert item.imported_from_excel is True


def test_normal_maintenance_excel_stores_unmatched_device_snapshot(client, db):
    general = SubProject(project_id=1, name="General")
    general.site_assignments = [SubProjectSite(site_id=1)]
    db.add(general)
    db.commit()
    login(client, *LEADER_A)
    page = client.get("/general-maintenance")
    csrf = re.search(r'name="csrf_token" value="([^"]+)"', page.text).group(1)
    form_token = re.search(r'name="form_token" value="([^"]+)"', page.text).group(1)
    preview = client.post(
        "/data-entry/maintenance/device-import-preview",
        data={"csrf_token": csrf, "project_id": "1", "sub_project_id": str(general.id), "work_site_id": "1"},
        files={"device_file": ("completed.xlsx", _device_workbook(db, serial="MAINT-SNAPSHOT-001"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    import_payload = preview.json()
    response = client.post(
        "/general-maintenance/submit",
        data={
            "csrf_token": csrf,
            "form_token": form_token,
            "project_id": "1",
            "sub_project_id": str(general.id),
            "quotation_number": ensure_service_quotation("1"),
            "work_site_id": "1",
            "service_type_id": "1",
            "installed_device_id": "catalog:1",
            "result_0": "completed_successfully",
            "notes": "Completed normal maintenance.",
            "device_import_token": import_payload["token"],
        },
        files={"photos_0": ("proof.jpg", make_image(), "image/jpeg")},
    )
    assert response.status_code == 303
    db.expire_all()
    item = db.query(GeneralMaintenanceRecord).one().work_items[0]
    assert item.serial_number == "MAINT-SNAPSHOT-001"
    assert item.installed_device_id is None
    assert item.sim_type == "stc"
    assert item.location_name == "Gate 1"
    assert item.imported_from_excel is True


def test_one_installation_submission_keeps_cross_project_sites_under_one_record(client, db):
    first_sub = SubProject(project_id=1, name="Batch One")
    first_sub.site_assignments = [SubProjectSite(site_id=1)]
    second_sub = SubProject(project_id=3, name="Batch Two")
    second_sub.site_assignments = [SubProjectSite(site_id=2)]
    db.add_all([first_sub, second_sub])
    db.commit()
    login(client, *LEADER_A)
    page = client.get("/installations/submit")
    csrf = re.search(r'name="csrf_token" value="([^"]+)"', page.text).group(1)
    form_token = re.search(r'name="form_token" value="([^"]+)"', page.text).group(1)
    import_tokens = []
    for project_id, sub_project_id, site_id, serial, imei, iccid in (
        ("1", str(first_sub.id), "1", "BATCH-INSTALL-001", "111111111111111", "111111111111111111"),
        ("3", str(second_sub.id), "2", "BATCH-INSTALL-002", "222222222222222", "222222222222222222"),
    ):
        preview = client.post(
            "/data-entry/installation/device-import-preview",
            data={
                "csrf_token": csrf,
                "project_id": project_id,
                "sub_project_id": sub_project_id,
                "work_site_id": site_id,
            },
                files={"device_file": ("completed.xlsx", _device_workbook(db, serial=serial, imei=imei, iccid=iccid), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert preview.status_code == 200
        import_tokens.append(preview.json()["token"])
    response = client.post(
        "/installations/submit",
        data={
            "csrf_token": csrf,
            "form_token": form_token,
            "project_id": ["1", "3"],
            "sub_project_id": [str(first_sub.id), str(second_sub.id)],
            "work_site_id": ["1", "2"],
            "quotation_number": [ensure_service_quotation(1), ensure_service_quotation(3)],
            "item_scope_index": ["0", "1"],
            "service_type_id": ["1", "1"],
            "device_id": ["1", "1"],
            "serial_number": ["BATCH-INSTALL-001", "BATCH-INSTALL-002"],
            "warranty_start": ["2026-08-09", "2026-08-09"],
            "result_0": "completed_successfully",
            "result_1": "completed_successfully",
            "notes": ["First site installation.", "Second site installation."],
            "handover_notes": ["", ""],
            "device_import_token": import_tokens,
        },
        files=[
            ("photos_0", ("first.jpg", make_image(), "image/jpeg")),
            ("photos_1", ("second.jpg", make_image(), "image/jpeg")),
        ],
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert response.status_code == 201, response.text
    db.expire_all()
    records = db.query(InstallationRecord).order_by(InstallationRecord.id).all()
    assert len(records) == 1
    record = records[0]
    assert response.json()["record_numbers"] == [record.record_number]
    assert [(item.project_id, item.sub_project_id, item.work_site_id) for item in record.work_items] == [
        (1, first_sub.id, 1),
        (3, second_sub.id, 2),
    ]
    assert [item.scope_position for item in record.work_items] == [0, 1]
    assert all(item.imported_from_excel for item in record.work_items)
    assert all(len(item.photos) == 1 for item in record.work_items)
    report_response = client.post(
        "/reports/installation",
        data=_report_payload(
            client,
            record.id,
            record_ids=[str(record.id)],
            include_device_data="1",
        ),
    )
    assert report_response.status_code == 303
    report = db.query(ServiceReport).one()
    pdf = client.get(f"/reports/installation/{report.id}/pdf?include_device_data=true")
    pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf.content)).pages)
    assert pdf_text.count("Device Data") == 2
    assert "Tower A" in pdf_text
    assert "Tower B" in pdf_text


def test_report_record_filter_accepts_second_precision_from_and_to(client, db):
    login(client, *LEADER_A)
    assert submit_installation(client, serial_number="TIME-FILTER-001").status_code == 303
    record = db.query(InstallationRecord).one()
    displayed = to_display(record.submitted_at).replace(microsecond=0)
    exact = displayed.strftime("%Y-%m-%dT%H:%M:%S")
    included = client.get(
        "/reports/installation/new",
        params={"from_at": exact, "to_at": exact},
    )
    assert included.status_code == 200
    assert record.record_number in included.text
    excluded = client.get(
        "/reports/installation/new",
        params={"from_at": (displayed + timedelta(seconds=1)).strftime("%Y-%m-%dT%H:%M:%S")},
    )
    assert excluded.status_code == 200
    assert record.record_number not in excluded.text


def test_report_record_filter_accepts_full_or_partial_record_number(client, db):
    login(client, *LEADER_A)
    assert submit_installation(client, serial_number="RECORD-SEARCH-001").status_code == 303
    first = db.query(InstallationRecord).one()
    assert submit_installation(client, serial_number="RECORD-SEARCH-002").status_code == 303
    second = db.query(InstallationRecord).order_by(InstallationRecord.id.desc()).first()
    client.get("/dashboard")

    exact = client.get(
        "/reports/installation/new",
        params={"record_number": first.record_number},
    )
    assert exact.status_code == 200
    assert first.record_number in exact.text
    assert second.record_number not in exact.text
    partial = client.get(
        "/reports/installation/new",
        params={"record_number": first.record_number[-5:]},
    )
    assert partial.status_code == 200
    assert first.record_number in partial.text
    assert 'name="record_number"' in partial.text


def _cross_project_scopes(db):
    first_sub = SubProject(project_id=1, name="Service Batch One")
    first_sub.site_assignments = [SubProjectSite(site_id=1)]
    second_sub = SubProject(project_id=3, name="Service Batch Two")
    second_sub.site_assignments = [SubProjectSite(site_id=2)]
    db.add_all([first_sub, second_sub])
    db.commit()
    return first_sub, second_sub


def _multi_maintenance_payload(client, page_url, first_sub, second_sub):
    page = client.get(page_url)
    return {
        "csrf_token": re.search(r'name="csrf_token" value="([^"]+)"', page.text).group(1),
        "form_token": re.search(r'name="form_token" value="([^"]+)"', page.text).group(1),
        "project_id": ["1", "3"],
        "sub_project_id": [str(first_sub.id), str(second_sub.id)],
        "work_site_id": ["1", "2"],
        "quotation_number": [ensure_service_quotation(1), ensure_service_quotation(3)],
        "item_scope_index": ["0", "1"],
        "service_type_id": ["1", "1"],
        "installed_device_id": ["catalog:1", "catalog:1"],
        "result_0": "completed_successfully",
        "result_1": "completed_successfully",
        "notes": ["First site service.", "Second site service."],
        "issue_description": ["", ""],
        "recommendations": ["", ""],
    }


def test_one_preventive_submission_keeps_cross_project_sites_under_one_record(client, db):
    first_sub, second_sub = _cross_project_scopes(db)
    login(client, *LEADER_A)
    response = client.post(
        "/maintenance/submit",
        data=_multi_maintenance_payload(client, "/maintenance/submit", first_sub, second_sub),
        files=[
            ("photos_0", ("first.jpg", make_image(), "image/jpeg")),
            ("photos_1", ("second.jpg", make_image(), "image/jpeg")),
        ],
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert response.status_code == 201, response.text
    db.expire_all()
    records = db.query(MaintenanceRecord).order_by(MaintenanceRecord.id).all()
    assert len(records) == 1
    assert [(item.project_id, item.sub_project_id, item.work_site_id) for item in records[0].work_items] == [
        (1, first_sub.id, 1),
        (3, second_sub.id, 2),
    ]
    assert all(len(item.photos) == 1 for item in records[0].work_items)


def test_one_maintenance_submission_keeps_cross_project_sites_under_one_record(client, db):
    first_sub, second_sub = _cross_project_scopes(db)
    login(client, *LEADER_A)
    response = client.post(
        "/general-maintenance/submit",
        data=_multi_maintenance_payload(client, "/general-maintenance", first_sub, second_sub),
        files=[
            ("photos_0", ("first.jpg", make_image(), "image/jpeg")),
            ("photos_1", ("second.jpg", make_image(), "image/jpeg")),
        ],
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert response.status_code == 201, response.text
    db.expire_all()
    records = db.query(GeneralMaintenanceRecord).order_by(GeneralMaintenanceRecord.id).all()
    assert len(records) == 1
    assert [(item.project_id, item.sub_project_id, item.work_site_id) for item in records[0].work_items] == [
        (1, first_sub.id, 1),
        (3, second_sub.id, 2),
    ]
    assert all(len(item.photos) == 1 for item in records[0].work_items)
